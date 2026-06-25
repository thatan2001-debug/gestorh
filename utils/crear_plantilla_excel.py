"""
Genera la plantilla Base_Empleados.xlsx con encabezados, una fila de ejemplo
y formato básico (anchos de columna, encabezado con color).
Ejecutar una sola vez: python utils/crear_plantilla_excel.py
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNAS = [
    "Nombre", "Documento", "Cargo", "Salario", "Fecha ingreso",
    "Fecha retiro", "Tipo contrato", "Correo",
]

EJEMPLO = [
    {
        "Nombre": "Juan Pérez",
        "Documento": "1020304050",
        "Cargo": "Auxiliar Administrativo",
        "Salario": 1800000,
        "Fecha ingreso": "01/02/2024",
        "Fecha retiro": "",
        "Tipo contrato": "Indefinido",
        "Correo": "empleado@email.com",
    }
]


def crear_plantilla(ruta_salida="plantillas/Base_Empleados.xlsx"):
    df = pd.DataFrame(EJEMPLO, columns=COLUMNAS)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Empleados")
        ws = writer.sheets["Empleados"]

        # Encabezado con estilo
        header_fill = PatternFill(start_color="1F3B73", end_color="1F3B73", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, col_name in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Anchos de columna
        anchos = [22, 16, 24, 14, 16, 16, 16, 28]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        # Nota de instrucciones en una segunda hoja
        notas = pd.DataFrame({
            "Instrucciones": [
                "Llena una fila por cada empleado.",
                "Nombre, Documento, Cargo, Salario y Fecha ingreso son obligatorios.",
                "Fecha ingreso y Fecha retiro en formato dd/mm/aaaa.",
                "Deja 'Fecha retiro' vacía si el empleado sigue activo.",
                "Tipo contrato: Fijo, Indefinido, Obra o labor, Prestación de servicios.",
                "El Salario debe ser un número, sin puntos ni signos ($).",
            ]
        })
        notas.to_excel(writer, index=False, sheet_name="Instrucciones")
        ws2 = writer.sheets["Instrucciones"]
        ws2.column_dimensions["A"].width = 70

    print(f"Plantilla creada en: {ruta_salida}")


if __name__ == "__main__":
    crear_plantilla()
